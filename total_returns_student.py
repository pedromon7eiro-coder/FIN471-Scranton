## ============================================================
##  Daily Total Return Downloader
##  Source: Yahoo Finance
## ------------------------------------------------------------
##  Professor Pedro Monteiro
##  FIN 471 — University of Scranton
## ============================================================
##
##  WHAT THIS SCRIPT GENERATES
##  --------------------------
##  An Excel workbook (daily_total_returns.xlsx) with one sheet
##  per ticker, plus a combined Summary sheet. Each sheet contains:
##
##    • Date          — trading day (YYYY-MM-DD)
##    • Close         — actual raw closing price investors transact at
##    • Dividends     — cash dividend paid on that date (0 on most days)
##    • Total Return  — (Close_t + Dividends_t - Close_{t-1}) / Close_{t-1}
##                      true buy-and-hold daily total return, including
##                      any dividend received that day
##
##  The Summary sheet shows all tickers side-by-side for easy comparison.
##
##  TWO ADDITIONAL CHART SHEETS
##  ---------------------------
##  • Chart — Normalized Prices : All tickers rebased to 100 on day 1,
##    showing relative price appreciation over the period.
##  • Chart — Cumulative Return  : Cumulative total return index (starting
##    at 100), compounding daily total returns day-by-day (dividends
##    included). This is the true wealth-growth comparison.
##
##  NOTE ON PRICES
##  --------------
##  Prices are raw market closing prices (what an investor actually paid
##  or received), NOT backward-adjusted prices. Dividends are captured
##  explicitly in the Dividends column and included in the Total Return
##  formula above, so no return information is lost.
##
##  NOTE ON THE FIRST ROW
##  ---------------------
##  The first trading day in the date range has no prior closing price,
##  so its Total Return is left blank (N/A). All subsequent rows are complete.
##
## ============================================================
##  INSTRUCTIONS: Edit only the three parameters below.
##  Do not modify anything beneath the divider line.
## ============================================================

# --------------------------------------------------------------
#  PARAMETER 1: Tickers
#  Format: "TICKER": "Description"
#
#  Reference:
#    Equities   -- SPY (S&P 500), IWM (Russell 2000), QQQ (Nasdaq),
#                  DIA (Dow Jones), VTI (Total US Market)
#    High Yield -- HYG, JNK
#    Corp Bond  -- LQD (Investment Grade)
#    Treasuries -- TLT (20+ yr), IEF (7-10 yr), SHY (1-3 yr)
#    Intl       -- EFA (Developed Markets), EEM (Emerging Markets)
#    REIT       -- VNQ
# --------------------------------------------------------------
TICKERS = {
    "SPY": "S&P 500",
    "IWM": "Russell 2000",
    "HYG": "High Yield Bond",
    "TLT": "Treasury Bond (20+ yr)",
}
# --------------------------------------------------------------
#  PARAMETER 2: Date Range
# --------------------------------------------------------------
START_DATE = "2020-01-01"   # YYYY-MM-DD
END_DATE   = "today"        # "today" or a specific date: "2026-05-09"
# --------------------------------------------------------------
#  PARAMETER 3: Output File Name
# --------------------------------------------------------------
OUTPUT_FILE = "daily_total_returns.xlsx"

## ============================================================
##  DO NOT EDIT BELOW THIS LINE
## ============================================================
!pip install yfinance openpyxl -q

import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import Series
import openpyxl.chart.series

# ── palette ────────────────────────────────────────────────────
NAVY      = "0D1B2A"
STEEL     = "1B3A5C"
GOLD      = "C9A84C"
WHITE     = "FFFFFF"
LIGHT_BG  = "F4F6FA"
ALT_ROW   = "EAF0FB"
BORDER_C  = "C5CFE0"
GREEN_POS = "1D7A4A"
RED_NEG   = "C0392B"
DIV_COLOR = "7B5EA7"

# Chart line colors (one per ticker, up to 10)
CHART_COLORS = [
    "2563EB",  # blue
    "DC2626",  # red
    "16A34A",  # green
    "D97706",  # amber
    "7C3AED",  # purple
    "0891B2",  # cyan
    "DB2777",  # pink
    "65A30D",  # lime
    "EA580C",  # orange
    "475569",  # slate
]

def side(color=BORDER_C, style="thin"):
    return Side(style=style, color=color)

THIN_BORDER = Border(left=side(), right=side(), top=side(), bottom=side())
HEADER_BORDER = Border(
    left=side(NAVY, "medium"), right=side(NAVY, "medium"),
    top=side(NAVY, "medium"), bottom=side(GOLD, "medium")
)

def title_font(sz=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=sz, bold=bold, color=color)

def body_font(sz=10, bold=False, color="000000"):
    return Font(name="Calibri", size=sz, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def right_align():
    return Alignment(horizontal="right", vertical="center")

# ── download data ───────────────────────────────────────────────
end = datetime.today().strftime("%Y-%m-%d") if END_DATE == "today" else END_DATE
ticker_data = {}

for ticker, name in TICKERS.items():
    print(f"Downloading {ticker} ({name})...")
    raw = yf.download(
        ticker,
        start=START_DATE,
        end=end,
        auto_adjust=False,
        actions=True,
        progress=False,
    )
    if raw.empty:
        print(f"  WARNING: No data found for {ticker}.")
        continue

    def get_col(df, col):
        if isinstance(df.columns, pd.MultiIndex):
            return df[(col, ticker)].squeeze()
        return df[col].squeeze()

    close     = get_col(raw, "Close").round(2)
    dividends = get_col(raw, "Dividends")
    total_ret = (close + dividends - close.shift(1)) / close.shift(1)

    df = pd.DataFrame({
        "Date":         close.index.strftime("%Y-%m-%d"),
        "Close":        close.values,
        "Dividends":    dividends.values,
        "Total_Return": total_ret.values,
    }).iloc[1:].reset_index(drop=True)   # drop first row (no prior price)

    ticker_data[ticker] = (name, df)

if not ticker_data:
    print("No data downloaded. Check your ticker symbols.")
    raise SystemExit

# ── build workbook ──────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

def write_sheet(wb, sheet_name, ticker, name, df):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    ws.merge_cells("A1:D2")
    banner = ws["A1"]
    banner.value = f"{ticker}  ·  {name}"
    banner.font  = Font(name="Calibri", size=16, bold=True, color=WHITE)
    banner.fill  = fill(NAVY)
    banner.alignment = center()

    ws.merge_cells("A3:D3")
    sub = ws["A3"]
    sub.value = f"Daily Total Return  |  {df['Date'].iloc[0]}  →  {df['Date'].iloc[-1]}  |  {len(df):,} trading days"
    sub.font  = Font(name="Calibri", size=10, italic=True, color="D0D8E8")
    sub.fill  = fill(STEEL)
    sub.alignment = center()

    headers = ["Date", "Close ($)", "Dividends ($)", "Total Return (%)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font      = title_font(10)
        cell.fill      = fill(STEEL)
        cell.alignment = center()
        cell.border    = HEADER_BORDER

    for row_idx, row in df.iterrows():
        excel_row = row_idx + 5
        bg = ALT_ROW if row_idx % 2 == 0 else LIGHT_BG

        c = ws.cell(row=excel_row, column=1, value=row["Date"])
        c.font = body_font(color="2C3E50"); c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

        c = ws.cell(row=excel_row, column=2, value=row["Close"])
        c.font = body_font(); c.fill = fill(bg)
        c.number_format = '#,##0.00'; c.alignment = right_align()
        c.border = THIN_BORDER

        div_val = row["Dividends"]
        c = ws.cell(row=excel_row, column=3, value=div_val if div_val > 0 else None)
        c.font = Font(name="Calibri", size=10, bold=div_val > 0, color=DIV_COLOR if div_val > 0 else "AAAAAA")
        c.fill = fill(bg)
        c.number_format = '#,##0.0000'; c.alignment = right_align()
        c.border = THIN_BORDER

        ret_val = row["Total_Return"]
        if pd.notna(ret_val):
            c = ws.cell(row=excel_row, column=4, value=ret_val)
            color = GREEN_POS if ret_val >= 0 else RED_NEG
            c.font = Font(name="Calibri", size=10, bold=True, color=color)
            c.number_format = '0.0000%'; c.alignment = right_align()
        else:
            c = ws.cell(row=excel_row, column=4, value="N/A")
            c.font = body_font(color="AAAAAA"); c.alignment = right_align()
        c.fill = fill(bg); c.border = THIN_BORDER

    for col, width in zip(["A","B","C","D"], [14, 13, 16, 18]):
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 20

    last_row = len(df) + 4
    ws.conditional_formatting.add(
        f"D5:D{last_row}",
        ColorScaleRule(
            start_type="min",  start_color="FFCCCC",
            mid_type="num",    mid_value=0, mid_color="FFFFFF",
            end_type="max",    end_color="CCFFDD",
        )
    )

    print(f"  ✓ Sheet '{sheet_name}' written — {len(df):,} rows")
    return ws

# ── individual ticker sheets ────────────────────────────────────
for ticker, (name, df) in ticker_data.items():
    write_sheet(wb, ticker, ticker, name, df)

# ── summary sheet ───────────────────────────────────────────────
ws = wb.create_sheet("Summary", 0)
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A5"

all_dates = sorted(set(
    date for _, df in ticker_data.values() for date in df["Date"]
))
tickers_list = list(ticker_data.keys())

num_cols = 1 + len(tickers_list) * 2
ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=num_cols)
b = ws.cell(row=1, column=1)
b.value = "FIN 471  ·  Daily Total Return Summary  ·  University of Scranton"
b.font  = Font(name="Calibri", size=15, bold=True, color=WHITE)
b.fill  = fill(NAVY); b.alignment = center()

ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
s = ws.cell(row=3, column=1)
s.value = f"Professor Pedro Monteiro  |  {all_dates[0]}  →  {all_dates[-1]}  |  Source: Yahoo Finance"
s.font  = Font(name="Calibri", size=10, italic=True, color="D0D8E8")
s.fill  = fill(STEEL); s.alignment = center()

ws.cell(row=4, column=1, value="Date").font = title_font(10)
ws["A4"].fill = fill(STEEL); ws["A4"].alignment = center()
ws["A4"].border = HEADER_BORDER

col = 2
ticker_col_map = {}
for ticker in tickers_list:
    name = ticker_data[ticker][0]
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
    cell = ws.cell(row=4, column=col, value=f"{ticker} — {name}")
    cell.font = title_font(10); cell.fill = fill(STEEL)
    cell.alignment = center(); cell.border = HEADER_BORDER
    ticker_col_map[ticker] = col
    col += 2

ws.cell(row=5, column=1, value="Date").font = title_font(9)
ws["A5"].fill = fill(STEEL); ws["A5"].alignment = center()
ws["A5"].border = HEADER_BORDER

for ticker in tickers_list:
    c = ticker_col_map[ticker]
    for offset, label in enumerate(["Close ($)", "Total Return (%)"]):
        cell = ws.cell(row=5, column=c+offset, value=label)
        cell.font = title_font(9); cell.fill = fill(STEEL)
        cell.alignment = center(); cell.border = HEADER_BORDER

ws.freeze_panes = "A6"

data_lookup = {
    ticker: df.set_index("Date") for ticker, (_, df) in ticker_data.items()
}

for row_idx, date in enumerate(all_dates):
    excel_row = row_idx + 6
    bg = ALT_ROW if row_idx % 2 == 0 else LIGHT_BG

    c = ws.cell(row=excel_row, column=1, value=date)
    c.font = body_font(color="2C3E50"); c.fill = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER

    for ticker in tickers_list:
        col = ticker_col_map[ticker]
        df_t = data_lookup[ticker]

        if date in df_t.index:
            row_data = df_t.loc[date]
            cl = ws.cell(row=excel_row, column=col, value=row_data["Close"])
            cl.font = body_font(); cl.fill = fill(bg)
            cl.number_format = '#,##0.00'; cl.alignment = right_align()
            cl.border = THIN_BORDER
            ret_val = row_data["Total_Return"]
            if pd.notna(ret_val):
                rc = ws.cell(row=excel_row, column=col+1, value=ret_val)
                color = GREEN_POS if ret_val >= 0 else RED_NEG
                rc.font = Font(name="Calibri", size=10, bold=True, color=color)
                rc.number_format = '0.0000%'
            else:
                rc = ws.cell(row=excel_row, column=col+1, value="N/A")
                rc.font = body_font(color="AAAAAA")
            rc.fill = fill(bg); rc.alignment = right_align(); rc.border = THIN_BORDER
        else:
            for offset in range(2):
                ec = ws.cell(row=excel_row, column=col+offset, value="—")
                ec.font = body_font(color="CCCCCC"); ec.fill = fill(bg)
                ec.alignment = center(); ec.border = THIN_BORDER

ws.column_dimensions["A"].width = 14
col = 2
for ticker in tickers_list:
    ws.column_dimensions[get_column_letter(col)].width   = 13
    ws.column_dimensions[get_column_letter(col+1)].width = 17
    col += 2

ws.row_dimensions[1].height = 28
ws.row_dimensions[3].height = 18
ws.row_dimensions[4].height = 20
ws.row_dimensions[5].height = 18

last_row = len(all_dates) + 5
for ticker in tickers_list:
    col = ticker_col_map[ticker] + 1
    col_letter = get_column_letter(col)
    ws.conditional_formatting.add(
        f"{col_letter}6:{col_letter}{last_row}",
        ColorScaleRule(
            start_type="min",  start_color="FFCCCC",
            mid_type="num",    mid_value=0, mid_color="FFFFFF",
            end_type="max",    end_color="CCFFDD",
        )
    )

print(f"  ✓ Summary sheet written — {len(all_dates):,} trading days × {len(tickers_list)} tickers")


# ── build aligned series ────────────────────────────────────────
common_dates = sorted(set(
    date for _, df in ticker_data.values() for date in df["Date"]
))

norm_series   = {}   # normalized price index (first close = 100)
cumret_series = {}   # cumulative total return index (start = 100)

for ticker, (name, df) in ticker_data.items():
    df_idx = df.set_index("Date")
    prices, cum_rets = [], []
    cum_index, first_close = 100.0, None

    for date in common_dates:
        if date in df_idx.index:
            row   = df_idx.loc[date]
            close = row["Close"]
            ret   = row["Total_Return"]
            if first_close is None:
                first_close = close
            prices.append(100.0 * close / first_close)
            if pd.notna(ret):
                cum_index *= (1 + ret)
            cum_rets.append(cum_index)
        else:
            prices.append(float("nan"))
            cum_rets.append(float("nan"))

    norm_series[ticker]   = prices
    cumret_series[ticker] = cum_rets


# ── compute annual & full-period total returns ──────────────────
def compute_returns_table(series_values_dict, all_dates):
    """
    Returns a dict:
      {ticker: {"full": float, "years": {year_str: float}}}
    All values are decimal (0.12 = 12%).

    Annual return = (last valid index value in year) / (last valid index value
    in prior year, or first valid value overall if it's the first year) - 1.
    This correctly captures every trading day's return within the calendar year.
    """
    date_index = {d: i for i, d in enumerate(all_dates)}
    years = sorted(set(d[:4] for d in all_dates))
    result = {}

    for ticker, values in series_values_dict.items():
        entry = {"years": {}}

        # Full period: series starts at 100 by construction, so
        # full return = last_valid / 100 - 1
        first_valid = next((v for v in values if not np.isnan(v)), None)
        last_valid  = next((v for v in reversed(values) if not np.isnan(v)), None)
        entry["full"] = (last_valid / first_valid - 1) if (first_valid and last_valid) else float("nan")

        # Build a map: year -> last valid index value in that year
        year_end_val = {}
        for yr in years:
            yr_vals = [values[date_index[d]] for d in all_dates if d.startswith(yr)]
            valid   = [v for v in yr_vals if not np.isnan(v)]
            year_end_val[yr] = valid[-1] if valid else float("nan")

        for i, yr in enumerate(years):
            end_val = year_end_val.get(yr, float("nan"))
            if np.isnan(end_val):
                entry["years"][yr] = float("nan")
                continue

            if i == 0:
                # First year: base is the very first valid value of the whole series
                base_val = first_valid
            else:
                # Subsequent years: base is last valid value of the prior year
                base_val = year_end_val.get(years[i - 1], float("nan"))

            if base_val is None or np.isnan(base_val) or base_val == 0:
                entry["years"][yr] = float("nan")
            else:
                entry["years"][yr] = end_val / base_val - 1

        result[ticker] = entry

    return result, years


norm_returns_table,   years_norm   = compute_returns_table(norm_series,   common_dates)
cumret_returns_table, years_cumret = compute_returns_table(cumret_series, common_dates)


# ── helper: write returns summary table below the chart ─────────
def write_returns_table(ws, start_row, tickers_list, returns_table, years,
                        chart_colors):
    """Writes a styled returns table starting at start_row, column A."""
    n_tickers = len(tickers_list)
    n_cols    = 2 + n_tickers   # Period | Ticker1 | Ticker2 | ...

    # Section header
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row,   end_column=n_cols)
    h = ws.cell(row=start_row, column=1,
                value="Total Return Summary  (Dividends Reinvested)")
    h.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    h.fill = fill(STEEL); h.alignment = center()
    ws.row_dimensions[start_row].height = 20
    start_row += 1

    # Sub-header: Period | Ticker names
    ws.cell(row=start_row, column=1, value="Period").font = title_font(9)
    ws["A" + str(start_row)].fill      = fill(NAVY)
    ws["A" + str(start_row)].alignment = center()
    ws["A" + str(start_row)].border    = HEADER_BORDER

    # blank spacer column
    sp = ws.cell(row=start_row, column=2, value="")
    sp.fill = fill(NAVY); sp.border = HEADER_BORDER

    for j, ticker in enumerate(tickers_list):
        name = ticker_data[ticker][0]
        col  = 3 + j
        c    = ws.cell(row=start_row, column=col, value=f"{ticker}\n{name}")
        c.font      = title_font(9)
        c.fill      = fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = HEADER_BORDER
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 2
    ws.row_dimensions[start_row].height = 30
    start_row += 1

    # Full period row
    row_data = [("Full Period", returns_table[t]["full"]) for t in tickers_list]
    _write_return_row(ws, start_row, "Full Period", tickers_list,
                      [returns_table[t]["full"] for t in tickers_list],
                      bg=LIGHT_BG, bold=True)
    start_row += 1

    # Annual rows
    for i, yr in enumerate(years):
        bg = ALT_ROW if i % 2 == 0 else LIGHT_BG
        _write_return_row(ws, start_row, str(yr), tickers_list,
                          [returns_table[t]["years"].get(yr, float("nan"))
                           for t in tickers_list],
                          bg=bg, bold=False)
        start_row += 1

    return start_row


def _write_return_row(ws, row, label, tickers_list, values, bg, bold):
    """Write one row of the returns table."""
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name="Calibri", size=10, bold=bold, color="2C3E50")
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER
    ws.row_dimensions[row].height = 18

    # spacer
    sp = ws.cell(row=row, column=2, value="")
    sp.fill = fill(bg); sp.border = THIN_BORDER

    for j, (ticker, val) in enumerate(zip(tickers_list, values)):
        col  = 3 + j
        cell = ws.cell(row=row, column=col)
        if not np.isnan(val):
            cell.value         = val
            cell.number_format = '0.00%'
            color = GREEN_POS if val >= 0 else RED_NEG
            cell.font = Font(name="Calibri", size=10, bold=bold, color=color)
        else:
            cell.value = "N/A"
            cell.font  = body_font(color="AAAAAA")
        cell.fill      = fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER


# ── helper: write a chart-only sheet with returns table ─────────
def write_chart_sheet(wb, sheet_title, chart_title, y_axis_title,
                      series_values_dict, all_dates,
                      returns_table, years):
    """
    Chart-only sheet (no data table visible to the user).
    Data is written to far-right hidden columns used only for the chart reference.
    A styled returns table is placed below the chart.
    """
    ws = wb.create_sheet(sheet_title)
    ws.sheet_view.showGridLines = False

    tickers_order = list(series_values_dict.keys())
    n_tickers     = len(tickers_order)

    # ── write chart data to hidden columns (far right, e.g. starting col 30) ──
    DATA_START_COL = 30   # col 29 reserved for year labels; cols A-X for chart/table
    date_col       = DATA_START_COL
    data_cols      = {}

    # header row for chart series labels (row 1)
    ws.cell(row=1, column=date_col, value="Date")
    for i, ticker in enumerate(tickers_order):
        col = DATA_START_COL + 1 + i
        name = ticker_data[ticker][0]
        ws.cell(row=1, column=col, value=f"{ticker} — {name}")
        data_cols[ticker] = col
        ws.column_dimensions[get_column_letter(col)].width = 0.1   # visually hidden

    ws.column_dimensions[get_column_letter(date_col)].width = 0.1

    # data rows (row 2 onward)
    n_data_rows = len(all_dates)
    for r, date in enumerate(all_dates, 2):
        ws.cell(row=r, column=date_col, value=date)
        for ticker, values in series_values_dict.items():
            col = data_cols[ticker]
            val = values[r - 2]
            ws.cell(row=r, column=col,
                    value=val if not np.isnan(val) else None)

    last_data_row = 1 + n_data_rows

    # ── compute how many data rows fall in each calendar year ──
    # We'll keep date strings (not serial numbers) for the category axis.
    # tickLblSkip tells Excel to only label every Nth category.
    # We find the index of the first trading day of each year so we can
    # place a label exactly there, then set tickLblSkip accordingly.
    years_in_data = sorted(set(d[:4] for d in all_dates))
    # Index (1-based) of the first occurrence of each year
    first_idx_of_year = {}
    for idx, d in enumerate(all_dates, 1):
        yr = d[:4]
        if yr not in first_idx_of_year:
            first_idx_of_year[yr] = idx

    # ── build LineChart ────────────────────────────────────────
    chart = LineChart()
    chart.title           = None
    chart.style           = 10
    chart.legend.position = "b"
    chart.width           = 32
    chart.height          = 18

    # ── Y-axis: gridlines + labels every 10 index points ──────
    all_vals = [v for vals in series_values_dict.values()
                for v in vals if not np.isnan(v)]
    data_min = min(all_vals) if all_vals else 50
    data_max = max(all_vals) if all_vals else 300
    y_min = float(max(0, (int(data_min * 0.97) // 10) * 10))
    y_min = min(y_min, 90.0)
    y_max = float(((int(data_max * 1.03) // 10) + 1) * 10)

    chart.y_axis.title       = y_axis_title
    chart.y_axis.scaling.min = y_min
    chart.y_axis.scaling.max = y_max
    chart.y_axis.majorUnit   = 10.0
    chart.y_axis.minorUnit   = 5.0
    chart.y_axis.numFmt      = '0'
    chart.y_axis.tickLblPos  = "low"
    chart.y_axis.crosses     = "min"
    chart.y_axis.delete      = False

    # ── X-axis: category axis, label once per year ─────────────
    # tickLblSkip=N shows a label every N categories.
    # We approximate: total rows / number of years → rows per year.
    rows_per_year = max(1, round(n_data_rows / max(len(years_in_data), 1)))
    chart.x_axis.title       = None
    chart.x_axis.tickLblPos  = "low"
    chart.x_axis.crosses     = "min"
    chart.x_axis.delete      = False
    chart.x_axis.tickMarkSkip = rows_per_year
    chart.x_axis.tickLblSkip  = rows_per_year
    # numFmt on a category axis only applies when the categories are dates;
    # since ours are strings we write the year label into a parallel
    # "label" column that contains the year string only on the first
    # trading day of each year and blank otherwise.
    LABEL_COL = DATA_START_COL - 1   # one column to the left of the date col
    ws.column_dimensions[get_column_letter(LABEL_COL)].width = 0.1
    ws.cell(row=1, column=LABEL_COL, value="Year")
    for r, date_str in enumerate(all_dates, 2):
        yr  = date_str[:4]
        val = yr if first_idx_of_year.get(yr) == (r - 1) else ""
        ws.cell(row=r, column=LABEL_COL, value=val)

    # Use the year-label column as x-axis categories
    cats_ref = Reference(ws, min_col=LABEL_COL, min_row=2, max_row=last_data_row)

    # Add one series per ticker
    for i, ticker in enumerate(tickers_order):
        col      = data_cols[ticker]
        data_ref = Reference(ws, min_col=col, min_row=1, max_row=last_data_row)
        chart.add_data(data_ref, titles_from_data=True)
        s = chart.series[-1]
        hex_color = CHART_COLORS[i % len(CHART_COLORS)]
        s.graphicalProperties.line.solidFill = hex_color
        s.graphicalProperties.line.width     = 20000
        s.smooth = True

    chart.set_categories(cats_ref)

    # ── banner (rows 1-2, columns A-Z) ────────────────────────
    VISIBLE_COLS = 2 + n_tickers   # matches returns table width
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=max(VISIBLE_COLS, 6))
    b = ws.cell(row=1, column=1, value=chart_title)
    b.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    b.fill = fill(NAVY); b.alignment = center()

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(VISIBLE_COLS, 6))
    s = ws.cell(row=3, column=1,
                value=f"FIN 471  ·  Prof. Pedro Monteiro  ·  {all_dates[0]} → {all_dates[-1]}  ·  Source: Yahoo Finance")
    s.font = Font(name="Calibri", size=9, italic=True, color="D0D8E8")
    s.fill = fill(STEEL); s.alignment = center()
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 16

    # ── place chart at row 4 ───────────────────────────────────
    ws.add_chart(chart, "A4")

    # chart height ≈ 18 cm; Excel rows ~0.53 cm each → ~34 rows
    chart_rows = 34
    returns_start = 4 + chart_rows + 1   # one blank row gap

    # ── returns table ──────────────────────────────────────────
    write_returns_table(ws, returns_start, tickers_order,
                        returns_table, years, CHART_COLORS)

    # ── column A width for the table label column ──────────────
    ws.column_dimensions["A"].width = 18

    print(f"  ✓ Chart sheet '{sheet_title}' written")
    return ws


# ── Chart Sheet 1 : Normalized Prices ──────────────────────────
write_chart_sheet(
    wb,
    sheet_title   = "Chart — Normalized Prices",
    chart_title   = "Normalized Price Index (Base = 100 on First Trading Day)",
    y_axis_title  = "Index Level (Base = 100)",
    series_values_dict = norm_series,
    all_dates     = common_dates,
    returns_table = norm_returns_table,
    years         = years_norm,
)

# ── Chart Sheet 2 : Cumulative Total Return ─────────────────────
write_chart_sheet(
    wb,
    sheet_title   = "Chart — Cumulative Return",
    chart_title   = "Cumulative Total Return Index (Base = 100, Dividends Reinvested)",
    y_axis_title  = "Index Level (Base = 100)",
    series_values_dict = cumret_series,
    all_dates     = common_dates,
    returns_table = cumret_returns_table,
    years         = years_cumret,
)

# ── save ────────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"\nComplete. Workbook saved as '{OUTPUT_FILE}'.")
print(f"Sheets: Summary | {list(ticker_data.keys())} | Chart — Normalized Prices | Chart — Cumulative Return")

from google.colab import files
files.download(OUTPUT_FILE)
